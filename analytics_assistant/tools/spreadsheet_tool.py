from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path
import re
from statistics import median, mean
from typing import Any

from analytics_assistant.config import settings
from analytics_assistant.models import Source, ToolResponse


def _safe_file_path(file_name: str) -> Path:
    base = settings.spreadsheet_dir.resolve()
    path = (base / file_name).resolve()
    if base not in path.parents and path != base:
        raise ValueError("Spreadsheet path must stay inside the configured spreadsheet directory.")
    if not path.exists():
        raise FileNotFoundError(f"Spreadsheet not found: {path}")
    return path


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as csv_file:
        return list(csv.DictReader(csv_file))


def _read_xlsx(path: Path, sheet: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "XLSX analysis requires openpyxl. Install dependencies with `pip install -r requirements.txt`."
        ) from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[sheet] if sheet else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) if value is not None else f"column_{index}" for index, value in enumerate(rows[0], start=1)]
    return [
        {headers[index]: value for index, value in enumerate(row)}
        for row in rows[1:]
    ]


def _read_table(path: Path, sheet: str | None) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        return _read_csv(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx(path, sheet)
    raise ValueError("Supported spreadsheet formats are .csv, .xlsx, and .xlsm.")


def _to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None

    rating_match = re.match(r"^\s*(-?\d+(?:\.\d+)?)\s*/\s*\d+(?:\.\d+)?\s*$", text)
    if rating_match:
        return float(rating_match.group(1))

    text = re.sub(r"\([^)]*\)", "", text)
    text = text.replace("$", "").replace(",", "").replace("%", "").strip()
    multiplier = 1.0
    suffix_match = re.match(r"^(-?\d+(?:\.\d+)?)([KMB])$", text, re.IGNORECASE)
    if suffix_match:
        text = suffix_match.group(1)
        suffix = suffix_match.group(2).lower()
        multiplier = {"k": 1_000.0, "m": 1_000_000.0, "b": 1_000_000_000.0}[suffix]
    elif re.search(r"[A-Za-z]", text):
        return None

    try:
        return float(text) * multiplier
    except ValueError:
        return None


def _query_tokens(query: str) -> set[str]:
    stopwords = {
        "a",
        "an",
        "and",
        "are",
        "by",
        "did",
        "do",
        "does",
        "for",
        "how",
        "in",
        "is",
        "of",
        "or",
        "the",
        "to",
        "was",
        "were",
        "what",
        "when",
        "why",
    }
    return {
        token.lower()
        for token in re.findall(r"[A-Za-z0-9][A-Za-z0-9_-]*", query)
        if len(token) > 2 and token.lower() not in stopwords
    }


def _row_text(row: dict[str, Any]) -> str:
    return " ".join("" if value is None else str(value) for value in row.values()).lower()


def _row_tokens(row: dict[str, Any]) -> set[str]:
    return set(_query_tokens(_row_text(row)))


def _numeric_values(rows: list[dict[str, Any]], column: str) -> list[float]:
    return [
        number
        for row in rows
        if (number := _to_number(row.get(column))) is not None
    ]


def _profile_rows(rows: list[dict[str, Any]], columns: list[str]) -> dict[str, Any]:
    numeric_columns: dict[str, dict[str, float | int]] = {}
    text_columns: dict[str, dict[str, Any]] = {}
    nonempty_counts: dict[str, int] = {}

    for column in columns:
        nonempty = [row.get(column) for row in rows if row.get(column) not in (None, "")]
        nonempty_counts[column] = len(nonempty)
        values = _numeric_values(rows, column)
        if values and len(values) >= max(1, int(len(nonempty) * 0.4)):
            numeric_columns[column] = {
                "count": len(values),
                "min": min(values),
                "max": max(values),
                "mean": mean(values),
                "median": median(values),
            }
        else:
            frequencies: dict[str, int] = defaultdict(int)
            for value in nonempty[:1000]:
                frequencies[str(value)] += 1
            top_values = sorted(frequencies.items(), key=lambda item: item[1], reverse=True)[:5]
            text_columns[column] = {
                "count": len(nonempty),
                "sample_values": [str(value) for value in nonempty[:3]],
                "top_values": [
                    {"value": value, "count": count}
                    for value, count in top_values
                ],
            }

    return {
        "row_count": len(rows),
        "columns": columns,
        "nonempty_counts": nonempty_counts,
        "numeric_columns": numeric_columns,
        "text_columns": text_columns,
        "inferred_roles": _infer_column_roles(columns),
    }


def _infer_column_roles(columns: list[str]) -> dict[str, list[str]]:
    role_patterns = {
        "title": ["title", "name", "movie", "release group"],
        "genre": ["genre", "category"],
        "rating": ["rating", "score", "meta", "méta"],
        "votes": ["vote"],
        "budget": ["budget", "cost"],
        "revenue": ["revenue", "gross", "worldwide", "domestic", "foreign", "box_office"],
        "date": ["date", "year"],
        "outcome": ["outcome", "status", "result", "description"],
    }
    roles: dict[str, list[str]] = {role: [] for role in role_patterns}
    for column in columns:
        normalized = column.lower().replace("$", "").replace("_", " ")
        for role, patterns in role_patterns.items():
            if any(pattern in normalized for pattern in patterns):
                roles[role].append(column)
    return {role: matched for role, matched in roles.items() if matched}


def _pick_rank_column(columns: list[str], rows: list[dict[str, Any]], query: str | None) -> str | None:
    numeric_columns = [
        column
        for column in columns
        if _numeric_values(rows, column)
    ]
    if not numeric_columns:
        return None

    lowered = (query or "").lower()
    preferences: list[str] = []
    if any(term in lowered for term in ["fail", "failed", "underperform", "low", "worst"]):
        preferences.extend(["rating", "score", "gross", "worldwide", "revenue"])
    if any(term in lowered for term in ["budget", "cost", "expensive"]):
        preferences.extend(["budget"])
    if any(term in lowered for term in ["revenue", "gross", "box office", "worldwide", "domestic"]):
        preferences.extend(["worldwide", "gross", "revenue", "domestic"])
    if any(term in lowered for term in ["rating", "review", "audience"]):
        preferences.extend(["rating", "score", "vote"])

    for preference in preferences:
        for column in numeric_columns:
            if preference in column.lower():
                return column
    return numeric_columns[0]


def _pick_vote_column(columns: list[str]) -> str | None:
    for column in columns:
        if "vote" in column.lower():
            return column
    return None


def _should_require_votes(query: str | None, rank_by: str | None) -> bool:
    lowered = (query or "").lower()
    rank = (rank_by or "").lower()
    return "rating" in rank and any(
        term in lowered for term in ["fail", "failed", "underperform", "low", "worst", "rating"]
    )


def _sort_reverse(sort_order: str, query: str | None) -> bool:
    if sort_order == "desc":
        return True
    if sort_order == "asc":
        return False
    lowered = (query or "").lower()
    return not any(term in lowered for term in ["fail", "failed", "underperform", "low", "worst"])


def _compact_row(row: dict[str, Any], max_columns: int = 14) -> dict[str, Any]:
    preferred_terms = [
        "title",
        "release",
        "genre",
        "rating",
        "vote",
        "budget",
        "gross",
        "worldwide",
        "domestic",
        "foreign",
        "revenue",
        "outcome",
        "description",
        "year",
        "rank",
    ]
    selected: dict[str, Any] = {}
    for column, value in row.items():
        if value in (None, ""):
            continue
        if any(term in column.lower() for term in preferred_terms):
            selected[column] = value
        if len(selected) >= max_columns:
            break
    if selected:
        return selected
    return {column: value for column, value in list(row.items())[:max_columns]}


def list_spreadsheets() -> ToolResponse:
    settings.spreadsheet_dir.mkdir(parents=True, exist_ok=True)
    files = [
        path.name
        for path in sorted(settings.spreadsheet_dir.iterdir())
        if path.suffix.lower() in {".csv", ".xlsx", ".xlsm"}
    ]
    return ToolResponse(
        data={"files": files},
        sources=[],
        explainability={"directory": str(settings.spreadsheet_dir)},
    )


def analyze_spreadsheet(
    file_name: str,
    operation: str = "describe",
    sheet: str | None = None,
    group_by: str | None = None,
    metric: str | None = None,
    aggregation: str = "sum",
    query: str | None = None,
    max_rows: int = 10,
    rank_by: str | None = None,
    sort_order: str = "auto",
    min_vote_count: float | None = None,
) -> ToolResponse:
    path = _safe_file_path(file_name)
    rows = _read_table(path, sheet)
    columns = list(rows[0].keys()) if rows else []

    if operation == "describe":
        profile = _profile_rows(rows, columns)
        data: Any = {
            "row_count": profile["row_count"],
            "columns": profile["columns"],
            "numeric_columns": profile["numeric_columns"],
        }
    elif operation == "auto_profile":
        data = _profile_rows(rows, columns)
    elif operation == "group_by":
        if not group_by or not metric:
            raise ValueError("group_by operation requires group_by and metric.")
        grouped: dict[str, list[float]] = defaultdict(list)
        for row in rows:
            number = _to_number(row.get(metric))
            if number is not None:
                grouped[str(row.get(group_by))].append(number)
        if aggregation == "sum":
            data = [{group_by: key, f"{metric}_sum": sum(values)} for key, values in grouped.items()]
        elif aggregation == "avg":
            data = [{group_by: key, f"{metric}_avg": mean(values)} for key, values in grouped.items()]
        elif aggregation == "count":
            data = [{group_by: key, f"{metric}_count": len(values)} for key, values in grouped.items()]
        else:
            raise ValueError("Supported aggregations are sum, avg, and count.")
    elif operation == "search":
        if not query:
            raise ValueError("search operation requires query.")
        tokens = _query_tokens(query)
        if not tokens:
            raise ValueError("search query does not contain searchable terms.")

        scored_rows: list[tuple[int, dict[str, Any]]] = []
        for row in rows:
            row_tokens = _row_tokens(row)
            score = sum(1 for token in tokens if token in row_tokens)
            if score:
                scored_rows.append((score, row))

        scored_rows.sort(key=lambda item: item[0], reverse=True)
        data = {
            "query": query,
            "matched_rows": [
                {"score": score, "row": row}
                for score, row in scored_rows[:max_rows]
            ],
            "searched_columns": columns,
        }
    elif operation == "filter_and_rank":
        tokens = _query_tokens(query or "")
        candidate_rows: list[tuple[int, dict[str, Any]]] = []
        for row in rows:
            row_tokens = _row_tokens(row)
            score = sum(1 for token in tokens if token in row_tokens) if tokens else 0
            if score or not tokens:
                candidate_rows.append((score, row))

        selected_rank_by = rank_by or _pick_rank_column(columns, rows, query)
        if selected_rank_by and selected_rank_by not in columns:
            raise ValueError(f"rank_by column not found: {selected_rank_by}")

        selected_vote_column = _pick_vote_column(columns)
        vote_threshold = min_vote_count
        if vote_threshold is None and _should_require_votes(query, selected_rank_by):
            vote_threshold = 1.0

        reverse = _sort_reverse(sort_order, query)
        if selected_rank_by:
            candidate_rows = [
                (score, row)
                for score, row in candidate_rows
                if _to_number(row.get(selected_rank_by)) is not None
                and (
                    vote_threshold is None
                    or selected_vote_column is None
                    or (_to_number(row.get(selected_vote_column)) or 0) >= vote_threshold
                )
            ]
            candidate_rows.sort(
                key=lambda item: (_to_number(item[1].get(selected_rank_by)) or 0, item[0]),
                reverse=reverse,
            )
        else:
            candidate_rows.sort(key=lambda item: item[0], reverse=True)

        ranked_rows = []
        for score, row in candidate_rows[:max_rows]:
            compact = _compact_row(row)
            ranked_rows.append(
                {
                    "match_score": score,
                    "rank_by": selected_rank_by,
                    "rank_value": _to_number(row.get(selected_rank_by)) if selected_rank_by else None,
                    "vote_column": selected_vote_column,
                    "vote_count": _to_number(row.get(selected_vote_column)) if selected_vote_column else None,
                    "row": compact,
                }
            )

        data = {
            "query": query,
            "rank_by": selected_rank_by,
            "sort_order": "desc" if reverse else "asc",
            "min_vote_count": vote_threshold,
            "matched_row_count": len(candidate_rows),
            "ranked_rows": ranked_rows,
            "inferred_roles": _infer_column_roles(columns),
        }
    else:
        raise ValueError(
            "Supported operations are describe, auto_profile, group_by, search, and filter_and_rank."
        )

    return ToolResponse(
        data=data,
        sources=[
            Source(
                type="spreadsheet",
                name=file_name,
                details={
                    "sheet": sheet,
                    "operation": operation,
                    "columns_used": [column for column in [group_by, metric, rank_by] if column] or columns,
                    "query": query,
                    "rank_by": rank_by,
                    "sort_order": sort_order,
                    "min_vote_count": min_vote_count,
                },
            )
        ],
        explainability={"file_path": str(path), "rows_read": len(rows)},
    )
